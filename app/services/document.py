import openpyxl

def copy_sheet(source_wb_path, target_wb_path, source_sheet_name, target_sheet_name):
    # Load the source workbook
    source_wb = openpyxl.load_workbook(source_wb_path)
    
    # Load the target workbook
    target_wb = openpyxl.load_workbook(target_wb_path)
    
    # Get the source sheet
    source_sheet = source_wb[source_sheet_name]
    
    # Remove the existing sheet with the target name in the target workbook
    if target_sheet_name in target_wb.sheetnames:
        target_wb.remove(target_wb[target_sheet_name])
    
    # Create a new sheet in the target workbook with the target name
    target_sheet = target_wb.create_sheet(title=target_sheet_name)
    
    # Copy the cell data, formulas, and styles from the source sheet to the target sheet
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)
    
    # Save the changes to the target workbook
    target_wb.save(target_wb_path)

# Paths to the source and target Excel workbooks
source_workbook_paths = [
    "/home/mohits/UAE (copy)/uploads/output.xlsx",
    "/home/mohits/UAE (copy)/uploads/output1.xlsx",
    "/home/mohits/UAE (copy)/uploads/output2.xlsx"
]
target_workbook_path = "/home/mohits/UAE (copy)/app/assets/Output_Format.xlsx"

# Sheet names to copy and corresponding target sheet names
sheets_to_copy = ['Client TB', 'Pivot OI', 'Input AJE']
target_sheet_names = ['Client TB', 'Pivot OI', 'Input AJE']

# Loop through each source workbook and corresponding sheet names
for source_path, sheet_name, target_name in zip(source_workbook_paths, sheets_to_copy, target_sheet_names):
    # Call the function to copy the sheet
    copy_sheet(source_path, target_workbook_path, sheet_name, target_name)
    print(f"The sheet '{sheet_name}' has been copied from '{source_path}' to '{target_workbook_path}'.")

# Note: Make sure the order of 'sheets_to_copy' and 'target_sheet_names' corresponds correctly.
